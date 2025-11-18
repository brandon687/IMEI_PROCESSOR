#!/usr/bin/env python3
"""
GSM Fusion Web Interface - PRODUCTION HARDENED
Zero-downtime version with comprehensive error handling
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session, send_file, Response, stream_with_context
from dotenv import load_dotenv
from gsm_fusion_client import GSMFusionClient, GSMFusionAPIError
from database import get_database
from production_submission_system import ProductionSubmissionSystem, SubmissionResult
from supabase_storage import get_storage
from export_completed_orders import export_completed_orders_to_csv, export_all_orders_to_csv, list_exported_csvs
from imei_data_parser import IMEIDataParser
from typing import Dict
import os
import logging
import traceback
from functools import wraps
import time
import csv
import io
import openpyxl
from datetime import datetime
import threading
import re
import json

# Setup logging
logging.basicConfig(
    level=os.environ.get('LOG_LEVEL', 'INFO'),
    format='%(asctime)s - %(name)s - [%(levelname)s] - %(message)s'
)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()

app = Flask(__name__)
app.secret_key = os.urandom(24)

# Global state
_db_instance = None
_services_cache = None
_services_cache_time = 0
CACHE_DURATION = 300  # 5 minutes


def get_db_safe():
    """Get database with error handling - never crashes"""
    global _db_instance
    if _db_instance is None:
        try:
            _db_instance = get_database()
            logger.info("✓ Database connected successfully")
            return _db_instance
        except ValueError as e:
            logger.error(f"Database config error: {e}")
            return None
        except Exception as e:
            logger.error(f"Database connection failed: {e}")
            return None
    return _db_instance


def get_services_cached(max_age=300):
    """Get services with caching and error handling"""
    global _services_cache, _services_cache_time

    now = time.time()
    if _services_cache and (now - _services_cache_time < max_age):
        logger.info(f"Using cached services ({len(_services_cache)} items)")
        return _services_cache

    try:
        logger.info("Fetching fresh services from API...")
        client = GSMFusionClient(timeout=10)  # 10 second timeout
        services = client.get_imei_services()
        client.close()

        _services_cache = services
        _services_cache_time = now
        logger.info(f"✓ Fetched {len(services)} services successfully")
        return services

    except Exception as e:
        logger.error(f"Failed to fetch services: {e}")
        logger.error(traceback.format_exc())

        # Return cached data even if stale
        if _services_cache:
            logger.warning(f"Returning stale cache ({len(_services_cache)} items)")
            return _services_cache

        # Return empty list as last resort
        return []


def get_service_name_by_id(service_id):
    """Get service name from service_id using cached services"""
    services = get_services_cached()
    for service in services:
        if str(service.get('id')) == str(service_id):
            return service.get('name', '')
    return ''


def error_handler(f):
    """Decorator to catch all errors and return error page"""
    @wraps(f)
    def wrapper(*args, **kwargs):
        try:
            return f(*args, **kwargs)
        except GSMFusionAPIError as e:
            logger.error(f"API Error in {f.__name__}: {e}")
            return render_template('error.html', error=f"API Error: {str(e)}"), 500
        except Exception as e:
            logger.error(f"Error in {f.__name__}: {e}")
            logger.error(traceback.format_exc())
            return render_template('error.html', error=f"Application Error: {str(e)}"), 500
    return wrapper


@app.route('/health')
def health_check():
    """Comprehensive health check endpoint for monitoring and deployments"""
    health_status = {
        'status': 'healthy',
        'timestamp': time.time(),
        'version': os.environ.get('VERSION', 'unknown'),
        'environment': os.environ.get('RAILWAY_ENVIRONMENT', 'local'),
        'checks': {},
        'metrics': {}
    }

    # Check database
    try:
        db = get_db_safe()
        if db:
            # Test database connectivity with a simple query
            conn = db.conn
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM orders")
            order_count = cursor.fetchone()[0]

            health_status['checks']['database'] = {
                'status': 'connected',
                'orders': order_count
            }
            health_status['metrics']['total_orders'] = order_count
        else:
            health_status['checks']['database'] = {
                'status': 'not_configured',
                'message': 'Database not initialized'
            }
            health_status['status'] = 'degraded'
    except Exception as e:
        health_status['checks']['database'] = {
            'status': 'error',
            'message': str(e)
        }
        health_status['status'] = 'degraded'

    # Check API client connectivity
    try:
        start_time = time.time()
        services = get_services_cached(max_age=60)
        response_time = round((time.time() - start_time) * 1000, 2)

        health_status['checks']['api'] = {
            'status': 'ok',
            'services': len(services),
            'response_time_ms': response_time,
            'cache_age_seconds': int(time.time() - _services_cache_time) if _services_cache else 0
        }
        health_status['metrics']['api_response_time_ms'] = response_time
        health_status['metrics']['services_available'] = len(services)

        if len(services) == 0:
            health_status['status'] = 'degraded'
            health_status['checks']['api']['status'] = 'warning'
            health_status['checks']['api']['message'] = 'No services available'

    except Exception as e:
        health_status['checks']['api'] = {
            'status': 'error',
            'message': str(e)
        }
        health_status['status'] = 'degraded'

    # Check cache
    health_status['checks']['cache'] = {
        'status': 'active' if _services_cache else 'empty',
        'items': len(_services_cache) if _services_cache else 0,
        'age_seconds': int(time.time() - _services_cache_time) if _services_cache else None
    }

    # Check environment variables
    required_env_vars = ['GSM_FUSION_API_KEY', 'GSM_FUSION_USERNAME']
    missing_vars = [var for var in required_env_vars if not os.environ.get(var)]

    if missing_vars:
        health_status['checks']['environment'] = {
            'status': 'error',
            'missing_variables': missing_vars
        }
        health_status['status'] = 'unhealthy'
    else:
        health_status['checks']['environment'] = {
            'status': 'ok',
            'api_key_length': len(os.environ.get('GSM_FUSION_API_KEY', '')),
            'username': os.environ.get('GSM_FUSION_USERNAME', 'unknown')
        }

    # Overall status determination
    check_statuses = [
        check.get('status') if isinstance(check, dict) else check
        for check in health_status['checks'].values()
    ]

    if 'error' in check_statuses or health_status['status'] == 'unhealthy':
        health_status['status'] = 'unhealthy'
        status_code = 503
    elif 'warning' in check_statuses or health_status['status'] == 'degraded':
        health_status['status'] = 'degraded'
        status_code = 200  # Still operational
    else:
        health_status['status'] = 'healthy'
        status_code = 200

    return jsonify(health_status), status_code


@app.route('/api/debug')
def api_debug():
    """Deep diagnostic endpoint - shows EXACT API response"""
    if not os.environ.get('ENABLE_DEBUG_ENDPOINT'):
        return jsonify({'error': 'Debug endpoint disabled. Set ENABLE_DEBUG_ENDPOINT=1 to enable'}), 403

    diagnostic = {
        'timestamp': time.time(),
        'environment': {
            'API_KEY_SET': bool(os.environ.get('GSM_FUSION_API_KEY')),
            'API_KEY_LENGTH': len(os.environ.get('GSM_FUSION_API_KEY', '')),
            'USERNAME': os.environ.get('GSM_FUSION_USERNAME', 'NOT_SET'),
            'BASE_URL': os.environ.get('GSM_FUSION_BASE_URL', 'http://hammerfusion.com'),
        },
        'api_test': {}
    }

    try:
        logger.info("=== DIAGNOSTIC API TEST STARTING ===")
        client = GSMFusionClient(timeout=10)

        # Make raw request
        xml_response = client._make_request('imeiservices')

        diagnostic['api_test'] = {
            'success': True,
            'raw_xml_length': len(xml_response),
            'raw_xml_first_500': xml_response[:500],
            'raw_xml_last_500': xml_response[-500:] if len(xml_response) > 500 else xml_response,
            'full_xml': xml_response  # FULL response for analysis
        }

        # Try parsing
        try:
            parsed = client._parse_xml_response(xml_response)
            diagnostic['api_test']['parsed_successfully'] = True
            diagnostic['api_test']['parsed_type'] = str(type(parsed))
            diagnostic['api_test']['parsed_keys'] = list(parsed.keys()) if isinstance(parsed, dict) else 'NOT A DICT'
            diagnostic['api_test']['parsed_data'] = str(parsed)[:1000]  # First 1000 chars
        except Exception as parse_error:
            diagnostic['api_test']['parsed_successfully'] = False
            diagnostic['api_test']['parse_error'] = str(parse_error)

        client.close()

    except Exception as e:
        diagnostic['api_test'] = {
            'success': False,
            'error': str(e),
            'error_type': type(e).__name__
        }

    return jsonify(diagnostic)


@app.route('/api/status')
def api_status():
    """Real-time API status for status bar - lightweight and fast"""
    status = {
        'timestamp': time.time(),
        'services': {
            'gsm_fusion': {'status': 'unknown', 'message': '', 'response_time': None},
            'database': {'status': 'unknown', 'message': ''},
            'cache': {'status': 'unknown', 'message': ''}
        },
        'overall': 'checking'
    }

    # Check GSM Fusion API
    try:
        start = time.time()
        client = GSMFusionClient(timeout=5)
        # Quick test - just check if we can create client
        client.close()
        response_time = round((time.time() - start) * 1000, 2)  # ms

        # Check if we have services cached
        services = get_services_cached(max_age=300)
        service_count = len(services)

        if service_count > 0:
            status['services']['gsm_fusion'] = {
                'status': 'operational',
                'message': f'{service_count} services available',
                'response_time': response_time
            }
        else:
            status['services']['gsm_fusion'] = {
                'status': 'degraded',
                'message': 'API responding but 0 services returned',
                'response_time': response_time
            }
    except Exception as e:
        status['services']['gsm_fusion'] = {
            'status': 'outage',
            'message': f'API Error: {str(e)[:100]}',
            'response_time': None
        }

    # Check Database
    try:
        db = get_db_safe()
        if db:
            status['services']['database'] = {
                'status': 'operational',
                'message': 'Supabase connected'
            }
        else:
            status['services']['database'] = {
                'status': 'not_configured',
                'message': 'Database not configured'
            }
    except Exception as e:
        status['services']['database'] = {
            'status': 'outage',
            'message': f'DB Error: {str(e)[:100]}'
        }

    # Check Cache
    if _services_cache:
        cache_age = int(time.time() - _services_cache_time)
        status['services']['cache'] = {
            'status': 'operational',
            'message': f'{len(_services_cache)} services (age: {cache_age}s)'
        }
    else:
        status['services']['cache'] = {
            'status': 'empty',
            'message': 'No cached data'
        }

    # Determine overall status
    statuses = [s['status'] for s in status['services'].values()]
    if 'outage' in statuses:
        status['overall'] = 'outage'
    elif 'degraded' in statuses:
        status['overall'] = 'degraded'
    elif all(s in ['operational', 'not_configured'] for s in statuses):
        status['overall'] = 'operational'
    else:
        status['overall'] = 'unknown'

    return jsonify(status)


@app.route('/')
@error_handler
def index():
    """Home page - bulletproof version"""
    logger.info("INDEX route called")

    # Check database (non-blocking)
    db = get_db_safe()
    if db is None:
        logger.warning("Database not available, continuing without it")

    # Get services with caching and error handling
    services = get_services_cached()

    if not services:
        logger.warning("No services available")
        return render_template('error.html',
                             error="Unable to load services. Please try again later."), 503

    # Get popular services (first 20)
    popular_services = services[:20] if len(services) > 20 else services

    logger.info(f"✓ Rendering index with {len(popular_services)} services")

    return render_template('index.html',
                         services=popular_services,
                         total_services=len(services),
                         recent_orders=[])


@app.route('/services')
@error_handler
def services_page():
    """Full services list page"""
    logger.info("SERVICES route called")

    services = get_services_cached()

    if not services:
        return render_template('error.html',
                             error="Unable to load services. Please try again later."), 503

    # Filter by category if provided
    category = request.args.get('category')
    search = request.args.get('search', '').lower()

    if category:
        services = [s for s in services if s.category == category]

    if search:
        services = [s for s in services if search in s.title.lower() or search in s.category.lower()]

    # Get unique categories
    categories = list(set([s.category for s in services]))

    return render_template('services.html',
                         services=services,
                         categories=categories,
                         selected_category=category,
                         search=search)


# ==========================================
# SUBMISSION ROUTES
# ==========================================

@app.route('/submit', methods=['GET', 'POST'])
@error_handler
def submit():
    """Submit single or multiple IMEI orders"""
    logger.info("SUBMIT route called")

    if request.method == 'POST':
        imei_input = request.form.get('imei', '').strip()
        service_id = request.form.get('service_id', '').strip()
        force_recheck = request.form.get('force_recheck') == 'true'

        if not imei_input or not service_id:
            flash('IMEI and Service ID are required', 'error')
            return redirect(url_for('submit'))

        # Parse multiple IMEIs (one per line)
        imei_lines = imei_input.strip().split('\n')
        imeis = []
        for line in imei_lines:
            imei = line.strip()
            if imei:
                # Validate IMEI (15 digits)
                if not imei.isdigit() or len(imei) != 15:
                    flash(f'Invalid IMEI: {imei}. Must be 15 digits. Skipped.', 'warning')
                    continue
                imeis.append(imei)

        if not imeis:
            flash('No valid IMEIs found. Each IMEI must be 15 digits.', 'error')
            return redirect(url_for('submit'))

        # Submit orders using GSM Fusion client
        try:
            client = GSMFusionClient(timeout=30)
            result = client.place_imei_order(imeis, service_id, force_recheck=force_recheck)
            client.close()

            # Store in database
            db = get_db_safe()
            if db and result['orders']:
                service_name = get_service_name_by_id(service_id)
                logger.info(f"Storing {len(result['orders'])} orders in database")
                for i, order in enumerate(result['orders'], 1):
                    try:
                        logger.info(f"Inserting order {i}/{len(result['orders'])}: order_id={order['id']}, imei={order['imei']}")
                        db.insert_order({
                            'order_id': order['id'],
                            'imei': order['imei'],
                            'service_id': service_id,
                            'service_name': service_name,
                            'status': order.get('status', 'Pending')
                        })
                        logger.info(f"✓ Successfully inserted order {order['id']}")
                    except Exception as e:
                        logger.error(f"❌ DB insert failed for order {order.get('id', 'unknown')}: {e}")
                        import traceback
                        logger.error(traceback.format_exc())

            # Show summary
            successful = len(result['orders'])
            duplicates = len(result['duplicates'])
            errors = len(result['errors'])

            if successful > 0:
                flash(f'✅ Submitted {successful} order(s) successfully!', 'success')
                if duplicates > 0:
                    flash(f'⚠️ {duplicates} duplicate(s) skipped', 'warning')
                if errors > 0:
                    flash(f'❌ {errors} error(s) occurred', 'error')
                return redirect(url_for('history'))
            else:
                if duplicates > 0:
                    flash(f'All IMEIs are duplicates ({duplicates} total)', 'warning')
                if errors > 0:
                    flash(f'Submission failed: {errors} error(s)', 'error')
                    for error in result['errors'][:3]:  # Show first 3 errors
                        flash(f'Error: {error}', 'error')
                return redirect(url_for('submit'))

        except Exception as e:
            logger.error(f"Submission error: {e}")
            logger.error(traceback.format_exc())
            flash(f'Submission failed: {str(e)}', 'error')
            return redirect(url_for('submit'))

    # GET request - show form
    services = get_services_cached()
    if not services:
        return render_template('error.html', error="Unable to load services"), 503

    return render_template('submit.html', services=services)


@app.route('/submit-stream', methods=['POST'])
def submit_stream():
    """
    Server-Sent Events endpoint for progressive IMEI order submission.

    Streams real-time progress updates during order submission process:
    - IMEI validation
    - Duplicate checking
    - API submission
    - Database storage

    Returns:
        Response: Server-Sent Events stream with progress updates

    Event Types:
        - progress: Status update with percentage
        - error: Error occurred, submission failed
        - complete: Submission successful with results

    Example Event:
        data: {"type": "progress", "step": "validating", "message": "Validating IMEI...", "percent": 10}
    """
    def generate():
        """Generator function for SSE stream"""
        start_time = time.time()

        try:
            # Parse form data
            imei_input = request.form.get('imei', '').strip()
            service_id = request.form.get('service_id', '').strip()
            force_recheck = request.form.get('force_recheck') == 'true'

            logger.info(f"[SSE] Starting streaming submission for service_id={service_id}, force_recheck={force_recheck}")

            # Step 1: Validation (10%)
            yield f"data: {json.dumps({'type': 'progress', 'step': 'validating', 'message': 'Validating IMEI numbers...', 'percent': 10})}\n\n"
            time.sleep(0.1)  # Small delay for visual feedback

            if not imei_input or not service_id:
                error_msg = 'IMEI and Service ID are required'
                logger.error(f"[SSE] Validation failed: {error_msg}")
                yield f"data: {json.dumps({'type': 'error', 'message': error_msg})}\n\n"
                return

            # Parse multiple IMEIs (one per line)
            imei_lines = imei_input.strip().split('\n')
            imeis = []
            invalid_imeis = []

            for line in imei_lines:
                imei = line.strip()
                if imei:
                    # Validate IMEI (15 digits)
                    if not imei.isdigit() or len(imei) != 15:
                        invalid_imeis.append(imei)
                        logger.warning(f"[SSE] Invalid IMEI format: {imei}")
                        continue
                    imeis.append(imei)

            if invalid_imeis:
                yield f"data: {json.dumps({'type': 'progress', 'step': 'validating', 'message': f'Skipped {len(invalid_imeis)} invalid IMEI(s)', 'percent': 20, 'warning': True})}\n\n"
                time.sleep(0.2)

            if not imeis:
                error_msg = 'No valid IMEIs found. Each IMEI must be 15 digits.'
                logger.error(f"[SSE] No valid IMEIs: {error_msg}")
                yield f"data: {json.dumps({'type': 'error', 'message': error_msg})}\n\n"
                return

            logger.info(f"[SSE] Validated {len(imeis)} IMEI(s)")
            yield f"data: {json.dumps({'type': 'progress', 'step': 'validated', 'message': f'Validated {len(imeis)} IMEI(s) successfully', 'percent': 25})}\n\n"
            time.sleep(0.1)

            # Step 2: Database duplicate check (30%)
            yield f"data: {json.dumps({'type': 'progress', 'step': 'checking_duplicates', 'message': 'Checking for existing orders...', 'percent': 30})}\n\n"

            db = get_db_safe()
            existing_count = 0
            if db and not force_recheck:
                try:
                    for imei in imeis:
                        existing_orders = db.get_orders_by_imei(imei)
                        if existing_orders:
                            existing_count += 1

                    if existing_count > 0:
                        logger.info(f"[SSE] Found {existing_count} existing order(s) in database")
                        yield f"data: {json.dumps({'type': 'progress', 'step': 'checking_duplicates', 'message': f'Found {existing_count} existing order(s) in database', 'percent': 35, 'warning': True})}\n\n"
                        time.sleep(0.2)
                except Exception as e:
                    logger.warning(f"[SSE] Database duplicate check failed: {e}")

            # Step 3: API submission (40-70%)
            yield f"data: {json.dumps({'type': 'progress', 'step': 'submitting', 'message': f'Submitting {len(imeis)} IMEI(s) to GSM Fusion API...', 'percent': 40})}\n\n"

            api_start = time.time()
            client = GSMFusionClient(timeout=30)

            try:
                result = client.place_imei_order(imeis, service_id, force_recheck=force_recheck)
                api_duration = time.time() - api_start

                logger.info(f"[SSE] API call completed in {api_duration:.2f}s - {len(result['orders'])} successful, {len(result['duplicates'])} duplicates, {len(result['errors'])} errors")

                yield f"data: {json.dumps({'type': 'progress', 'step': 'submitted', 'message': f'API responded in {api_duration:.2f}s', 'percent': 70})}\n\n"
                time.sleep(0.1)

            except Exception as e:
                api_duration = time.time() - api_start
                error_msg = f"API submission failed: {str(e)}"
                logger.error(f"[SSE] {error_msg} (after {api_duration:.2f}s)")
                yield f"data: {json.dumps({'type': 'error', 'message': error_msg, 'duration': api_duration})}\n\n"
                client.close()
                return
            finally:
                client.close()

            # Step 4: Database storage (80%)
            yield f"data: {json.dumps({'type': 'progress', 'step': 'saving', 'message': 'Saving orders to database...', 'percent': 80})}\n\n"

            saved_count = 0
            if db and result['orders']:
                for order in result['orders']:
                    try:
                        db.insert_order({
                            'order_id': order['id'],
                            'imei': order['imei'],
                            'service_id': service_id,
                            'status': order.get('status', 'Pending')
                        })
                        saved_count += 1
                    except Exception as e:
                        logger.warning(f"[SSE] DB insert failed for order {order['id']}: {e}")

                logger.info(f"[SSE] Saved {saved_count}/{len(result['orders'])} order(s) to database")

            yield f"data: {json.dumps({'type': 'progress', 'step': 'saved', 'message': f'Saved {saved_count} order(s) to database', 'percent': 90})}\n\n"
            time.sleep(0.1)

            # Step 5: Complete (100%)
            total_duration = time.time() - start_time
            successful = len(result['orders'])
            duplicates = len(result['duplicates'])
            errors = len(result['errors'])

            completion_data = {
                'type': 'complete',
                'message': f'Successfully submitted {successful} order(s)!',
                'percent': 100,
                'stats': {
                    'successful': successful,
                    'duplicates': duplicates,
                    'errors': errors,
                    'total_imeis': len(imeis),
                    'duration': round(total_duration, 2),
                    'api_duration': round(api_duration, 2)
                },
                'redirect': url_for('history')
            }

            logger.info(f"[SSE] Submission complete in {total_duration:.2f}s: {successful} successful, {duplicates} duplicates, {errors} errors")
            yield f"data: {json.dumps(completion_data)}\n\n"

        except Exception as e:
            error_msg = f"Unexpected error: {str(e)}"
            logger.error(f"[SSE] {error_msg}")
            logger.error(traceback.format_exc())
            yield f"data: {json.dumps({'type': 'error', 'message': error_msg})}\n\n"

    # Return SSE response with proper headers
    return Response(
        stream_with_context(generate()),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no',  # Disable nginx buffering
            'Connection': 'keep-alive'
        }
    )


@app.route('/batch', methods=['GET', 'POST'])
@error_handler
def batch_upload():
    """Batch CSV/Excel upload"""
    logger.info("BATCH route called")

    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file uploaded', 'error')
            return redirect(url_for('batch_upload'))

        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(url_for('batch_upload'))

        service_id = request.form.get('service_id', '').strip()
        if not service_id:
            flash('Service ID is required', 'error')
            return redirect(url_for('batch_upload'))

        try:
            # Read file data for both parsing and storage
            file_data = file.stream.read()

            # Upload file to Supabase Storage
            storage = get_storage()
            file_url = None
            if storage.available:
                try:
                    # Detect content type
                    if file.filename.endswith('.csv'):
                        content_type = 'text/csv'
                    elif file.filename.endswith(('.xlsx', '.xls')):
                        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    else:
                        content_type = 'application/octet-stream'

                    file_url = storage.upload_file(file.filename, file_data, content_type)
                    if file_url:
                        logger.info(f"✅ Uploaded file to Supabase Storage: {file_url}")
                    else:
                        logger.warning("File upload to Supabase Storage failed, continuing with processing")
                except Exception as e:
                    logger.warning(f"Supabase Storage upload failed: {e}, continuing with processing")
            else:
                logger.info("Supabase Storage not available, file not uploaded to cloud")

            # Parse file based on extension
            if file.filename.endswith('.csv'):
                # Read CSV
                stream = io.StringIO(file_data.decode("UTF8"), newline=None)
                csv_reader = csv.DictReader(stream)
                imeis = [row.get('imei', '').strip() for row in csv_reader if row.get('imei')]
            elif file.filename.endswith(('.xlsx', '.xls')):
                # Read Excel
                wb = openpyxl.load_workbook(io.BytesIO(file_data))
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                imei_col = headers.index('imei') if 'imei' in headers else 0
                imeis = [str(ws.cell(row, imei_col + 1).value).strip()
                        for row in range(2, ws.max_row + 1)
                        if ws.cell(row, imei_col + 1).value]
            else:
                flash('Invalid file format. Use CSV or Excel.', 'error')
                return redirect(url_for('batch_upload'))

            # Validate IMEIs
            valid_imeis = [imei for imei in imeis if imei.isdigit() and len(imei) == 15]

            if not valid_imeis:
                flash('No valid IMEIs found in file', 'error')
                return redirect(url_for('batch_upload'))

            flash(f'Processing {len(valid_imeis)} IMEIs from file...', 'info')

            # Submit batch
            client = GSMFusionClient(timeout=30)
            result = client.place_imei_order(valid_imeis, service_id)
            client.close()

            # Store in database
            db = get_db_safe()
            if db and result['orders']:
                service_name = get_service_name_by_id(service_id)
                logger.info(f"Storing {len(result['orders'])} orders in database")
                for i, order in enumerate(result['orders'], 1):
                    try:
                        logger.info(f"Inserting order {i}/{len(result['orders'])}: order_id={order['id']}, imei={order['imei']}")
                        db.insert_order({
                            'order_id': order['id'],
                            'imei': order['imei'],
                            'service_id': service_id,
                            'service_name': service_name,
                            'status': order.get('status', 'Pending')
                        })
                        logger.info(f"✓ Successfully inserted order {order['id']}")
                    except Exception as e:
                        logger.error(f"❌ DB insert failed for order {order.get('id', 'unknown')}: {e}")
                        import traceback
                        logger.error(traceback.format_exc())

            successful = len(result['orders'])
            duplicates = len(result['duplicates'])
            errors = len(result['errors'])

            # Record import history with file URL
            if db:
                db.record_batch_import(
                    filename=file.filename,
                    rows_imported=successful,
                    rows_skipped=duplicates + errors,
                    file_url=file_url
                )

            flash(f'Batch processed: {successful} successful, {duplicates} duplicates, {errors} errors', 'success')
            return redirect(url_for('history'))

        except Exception as e:
            logger.error(f"Batch upload error: {e}")
            logger.error(traceback.format_exc())
            flash(f'Batch upload failed: {str(e)}', 'error')
            return redirect(url_for('batch_upload'))

    # GET - show form
    services = get_services_cached()
    return render_template('batch.html', services=services)


# ==========================================
# HISTORY & ORDER ROUTES
# ==========================================

@app.route('/history')
@error_handler
def history():
    """View order history"""
    logger.info("HISTORY route called")

    search_imei = request.args.get('imei', '').strip()

    db = get_db_safe()
    if not db:
        flash('Database not available', 'warning')
        return render_template('history.html', orders=[], search_query='', search_count=0)

    try:
        if search_imei:
            # Parse multiple IMEIs
            imeis = [line.strip() for line in search_imei.split('\n')
                    if line.strip() and line.strip().isdigit() and len(line.strip()) == 15]

            if len(imeis) == 1:
                orders = db.search_orders_by_imei(imeis[0])
            elif len(imeis) > 1:
                # Search multiple IMEIs
                all_orders = []
                for imei in imeis:
                    all_orders.extend(db.search_orders_by_imei(imei))
                orders = all_orders
            else:
                orders = []
                flash('No valid IMEIs found', 'warning')

            # Calculate search count for template
            search_count = len(imeis)
        else:
            orders = db.get_recent_orders(limit=100)
            search_count = 0

        return render_template('history.html',
                             orders=orders,
                             search_query=search_imei,
                             search_count=search_count)

    except Exception as e:
        logger.error(f"History error: {e}")
        logger.error(traceback.format_exc())
        flash(f'Error loading history: {str(e)}', 'error')
        return render_template('history.html', orders=[], search_query='', search_count=0)


@app.route('/history/sync', methods=['GET'])
@error_handler
def sync_orders():
    """Sync pending orders with API to update their status"""
    logger.info("SYNC route called")

    db = get_db_safe()
    if not db:
        flash('Database not available', 'error')
        return redirect(url_for('history'))

    try:
        # Get all pending orders
        conn = db.conn
        cursor = conn.cursor()
        cursor.execute("SELECT order_id FROM orders WHERE status IN ('Pending', 'In Process', '1', '4')")
        pending_orders = cursor.fetchall()

        if not pending_orders:
            flash('No pending orders to sync', 'info')
            return redirect(url_for('history'))

        # Collect order IDs
        order_ids = [row[0] for row in pending_orders]
        logger.info(f"Syncing {len(order_ids)} pending orders")

        # Fetch status from API (batch)
        client = GSMFusionClient(timeout=30)
        updated_count = 0

        try:
            # API accepts comma-separated order IDs for batch lookup
            order_ids_str = ','.join(order_ids)
            orders = client.get_imei_orders(order_ids_str)

            # Update database with new status
            for order in orders:
                try:
                    cursor.execute("""
                        UPDATE orders
                        SET status = ?,
                            carrier = ?,
                            model = ?,
                            simlock = ?,
                            fmi = ?,
                            result_code = ?,
                            result_code_display = ?,
                            updated_at = CURRENT_TIMESTAMP
                        WHERE order_id = ?
                    """, (
                        order.status,
                        order.carrier or '',
                        order.model or '',
                        order.simlock or '',
                        order.fmi or '',
                        order.result_code or '',
                        order.result_code_display or '',
                        order.id
                    ))
                    updated_count += 1
                except Exception as e:
                    logger.warning(f"Failed to update order {order.id}: {e}")

            conn.commit()
            flash(f'✅ Synced {updated_count} orders successfully', 'success')

        except Exception as e:
            logger.error(f"API sync failed: {e}")
            flash(f'Sync failed: {str(e)}', 'error')
        finally:
            client.close()

        return redirect(url_for('history'))

    except Exception as e:
        logger.error(f"Sync error: {e}")
        logger.error(traceback.format_exc())
        flash(f'Sync failed: {str(e)}', 'error')
        return redirect(url_for('history'))


def sync_and_parse_orders(parse_enabled: bool = True) -> Dict:
    """
    Helper function to sync pending orders with API and parse result codes

    Args:
        parse_enabled: Whether to parse the result_code field (default: True)

    Returns:
        Dictionary with detailed stats and results
    """
    start_time = time.time()

    result = {
        'success': False,
        'stats': {
            'total_pending': 0,
            'synced': 0,
            'completed': 0,
            'parsed': 0,
            'parse_failures': 0
        },
        'sample_parsed': [],
        'errors': [],
        'duration': 0
    }

    try:
        logger.info("=== MANUAL SYNC & PARSE STARTED ===")
        logger.info(f"Parse enabled: {parse_enabled}")

        # Get database
        db = get_db_safe()
        if not db:
            result['errors'].append('Database not available')
            return result

        # Get all pending orders
        conn = db.conn
        cursor = conn.cursor()
        cursor.execute("SELECT order_id FROM orders WHERE status IN ('Pending', 'In Process', '1', '4')")
        pending_orders = cursor.fetchall()

        result['stats']['total_pending'] = len(pending_orders)
        logger.info(f"Found {len(pending_orders)} pending orders to sync")

        if not pending_orders:
            result['success'] = True
            result['duration'] = time.time() - start_time
            return result

        # Collect order IDs
        order_ids = [row[0] for row in pending_orders]

        # Initialize parser if enabled
        parser = IMEIDataParser() if parse_enabled else None

        # Fetch status from API (batch)
        logger.info("Calling GSM Fusion API to sync orders...")
        client = GSMFusionClient(timeout=60)

        try:
            # API accepts comma-separated order IDs for batch lookup
            order_ids_str = ','.join(order_ids)
            api_start = time.time()
            orders = client.get_imei_orders(order_ids_str)
            api_duration = time.time() - api_start

            logger.info(f"API returned {len(orders)} orders in {api_duration:.2f}s")
            result['stats']['synced'] = len(orders)

            # Update database with new status and parse CODE field
            for i, order in enumerate(orders, 1):
                try:
                    logger.info(f"Processing order {i}/{len(orders)}: {order.id}")

                    # Parse the result_code if enabled and order is completed
                    parsed_data = {}
                    if parse_enabled and order.code and order.status == 'Completed':
                        logger.info(f"Parsing CODE field for order {order.id}")
                        logger.debug(f"Raw CODE: {order.code[:200]}...")

                        try:
                            parsed = parser.parse(order.code)
                            parsed_dict = parsed.to_dict()

                            # Map parser field names to database field names
                            parsed_data = {
                                'carrier': parsed_dict.get('carrier'),
                                'model': parsed_dict.get('model'),
                                'simlock': parsed_dict.get('simlock'),
                                'fmi': parsed_dict.get('find_my_iphone'),
                                'imei2': parsed_dict.get('imei2_number'),
                                'serial_number': parsed_dict.get('serial_number'),
                                'meid': parsed_dict.get('meid_number'),
                                'gsma_status': parsed_dict.get('current_gsma_status'),
                                'purchase_date': parsed_dict.get('estimated_purchase_date'),
                                'applecare': parsed_dict.get('applecare_eligible'),
                                'tether_policy': parsed_dict.get('next_tether_policy')
                            }

                            # Remove None values
                            parsed_data = {k: v for k, v in parsed_data.items() if v is not None}

                            logger.info(f"✓ Parsed {len(parsed_data)} fields: {list(parsed_data.keys())}")
                            result['stats']['parsed'] += 1

                            # Add to sample (first 3)
                            if len(result['sample_parsed']) < 3:
                                result['sample_parsed'].append({
                                    'order_id': order.id,
                                    'imei': order.imei,
                                    'parsed_fields': parsed_data
                                })

                        except Exception as parse_error:
                            logger.warning(f"Failed to parse CODE for order {order.id}: {parse_error}")
                            result['stats']['parse_failures'] += 1

                    # Update database
                    if parsed_data:
                        # Update with parsed data
                        db.update_order_status(
                            order_id=order.id,
                            status=order.status,
                            code=order.code,
                            code_display=order.result_code_display,
                            service_name=order.package,
                            result_data=parsed_data
                        )
                        logger.info(f"✓ Updated order {order.id} with parsed data")
                    else:
                        # Simple status update
                        db.update_order_status(
                            order_id=order.id,
                            status=order.status,
                            code=order.code,
                            code_display=order.result_code_display,
                            service_name=order.package
                        )
                        logger.info(f"✓ Updated order {order.id} status")

                    # Count completed orders
                    if order.status == 'Completed':
                        result['stats']['completed'] += 1

                except Exception as e:
                    error_msg = f"Failed to update order {order.id}: {str(e)}"
                    logger.error(error_msg)
                    result['errors'].append(error_msg)

            result['success'] = True
            logger.info(f"✓ Manual sync completed: {result['stats']['synced']} synced, {result['stats']['parsed']} parsed")

        except Exception as e:
            error_msg = f"API sync failed: {str(e)}"
            logger.error(error_msg)
            logger.error(traceback.format_exc())
            result['errors'].append(error_msg)
        finally:
            client.close()

    except Exception as e:
        error_msg = f"Sync error: {str(e)}"
        logger.error(error_msg)
        logger.error(traceback.format_exc())
        result['errors'].append(error_msg)

    result['duration'] = round(time.time() - start_time, 2)
    logger.info(f"=== MANUAL SYNC & PARSE FINISHED in {result['duration']}s ===")

    return result


@app.route('/manual-sync', methods=['POST'])
def manual_sync():
    """
    Manual sync endpoint with integrated IMEI data parser

    Accepts optional 'parse' parameter (default: true)
    Returns detailed results with debugging info
    """
    logger.info("MANUAL-SYNC route called")

    try:
        # Get parse parameter (default: true)
        parse_enabled = request.form.get('parse', 'true').lower() in ['true', '1', 'yes']

        logger.info(f"Starting manual sync with parse={parse_enabled}")

        # Call helper function
        result = sync_and_parse_orders(parse_enabled=parse_enabled)

        # Return JSON response
        return jsonify(result), 200 if result['success'] else 500

    except Exception as e:
        logger.error(f"Manual sync error: {e}")
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e),
            'stats': {},
            'sample_parsed': [],
            'errors': [str(e)],
            'duration': 0
        }), 500


@app.route('/status/<order_id>')
@error_handler
def order_status(order_id):
    """Check order status"""
    logger.info(f"STATUS route called for order: {order_id}")

    db = get_db_safe()

    # Try database first
    if db:
        try:
            # Search by order_id or IMEI
            conn = db.conn
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM orders WHERE order_id = ? OR imei = ? LIMIT 1", (order_id, order_id))
            row = cursor.fetchone()

            if row:
                columns = [desc[0] for desc in cursor.description]
                order = dict(zip(columns, row))
                return render_template('status.html', order=order)
        except Exception as e:
            logger.warning(f"DB lookup failed: {e}")

    # Fallback to API
    try:
        client = GSMFusionClient(timeout=10)
        orders = client.get_imei_orders(order_id)
        client.close()

        if not orders:
            flash('Order not found', 'error')
            return redirect(url_for('index'))

        # Convert IMEIOrder to dict for template
        order = {
            'order_id': orders[0].id,
            'imei': orders[0].imei,
            'service_name': orders[0].package,
            'status': orders[0].status,
            'code': orders[0].code,
            'order_date': orders[0].requested_at
        }

        return render_template('status.html', order=order)
    except Exception as e:
        logger.error(f"Order status error: {e}")
        logger.error(traceback.format_exc())
        return render_template('error.html', error=f"Unable to fetch order status: {str(e)}")


# ==========================================
# SERVICE & DATABASE ROUTES
# ==========================================

@app.route('/service/<service_id>')
@error_handler
def service_detail(service_id):
    """View service details"""
    logger.info(f"SERVICE DETAIL route called for: {service_id}")

    services = get_services_cached()

    # Find service
    service = next((s for s in services if s.package_id == service_id), None)

    if not service:
        flash('Service not found', 'error')
        return redirect(url_for('services_page'))

    return render_template('service_detail.html', service=service)


@app.route('/database')
@error_handler
def database_view():
    """View database statistics"""
    logger.info("DATABASE route called")

    db = get_db_safe()
    if not db:
        flash('Database not available', 'error')
        return redirect(url_for('index'))

    try:
        # Get recent orders
        orders = db.get_recent_orders(limit=50)

        # Calculate simple stats
        total_orders = len(orders)
        completed = len([o for o in orders if o.get('status', '').lower() == 'completed'])
        pending = len([o for o in orders if o.get('status', '').lower() in ['pending', 'in process']])

        # Calculate total credits from all orders
        total_credits = sum(float(o.get('credits', 0)) for o in orders if o.get('credits'))

        # Count orders today
        from datetime import datetime, timedelta
        today = datetime.now().date()
        orders_today = 0
        for o in orders:
            order_date = o.get('order_date') or o.get('created_at', '')
            if order_date:
                try:
                    # Parse date (handles both datetime and date strings)
                    if isinstance(order_date, str):
                        order_date_obj = datetime.strptime(order_date.split()[0], '%Y-%m-%d').date()
                    else:
                        order_date_obj = order_date.date() if hasattr(order_date, 'date') else order_date

                    if order_date_obj == today:
                        orders_today += 1
                except:
                    pass

        # Group by status
        by_status = {}
        for o in orders:
            status = o.get('status', 'Unknown')
            by_status[status] = by_status.get(status, 0) + 1

        stats = {
            'total_orders': total_orders,
            'completed': completed,
            'pending': pending,
            'total_credits': total_credits,
            'orders_today': orders_today,
            'by_status': by_status
        }

        return render_template('database.html',
                             stats=stats,
                             orders=orders,
                             recent_orders=orders[:20])
    except Exception as e:
        logger.error(f"Database view error: {e}")
        logger.error(traceback.format_exc())
        flash(f'Database error: {str(e)}', 'error')
        return redirect(url_for('index'))


# ==========================================
# CSV EXPORT ROUTES
# ==========================================

@app.route('/export-completed', methods=['GET'])
@error_handler
def export_completed():
    """Export completed orders to CSV and upload to Supabase Storage"""
    logger.info("EXPORT-COMPLETED route called")

    try:
        # Export completed orders
        csv_url = export_completed_orders_to_csv(status_filter='Completed')

        if csv_url:
            flash(f'✅ Exported completed orders to CSV: {csv_url}', 'success')
        else:
            flash('Failed to export completed orders. Check logs for details.', 'error')

        return redirect(url_for('database_view'))

    except Exception as e:
        logger.error(f"Export error: {e}")
        logger.error(traceback.format_exc())
        flash(f'Export failed: {str(e)}', 'error')
        return redirect(url_for('database_view'))


@app.route('/export-all', methods=['GET'])
@error_handler
def export_all():
    """Export all recent orders to CSV and upload to Supabase Storage"""
    logger.info("EXPORT-ALL route called")

    try:
        # Get limit from query parameter (default 10000)
        limit = int(request.args.get('limit', 10000))

        # Export all orders
        csv_url = export_all_orders_to_csv(limit=limit)

        if csv_url:
            flash(f'✅ Exported all orders to CSV: {csv_url}', 'success')
        else:
            flash('Failed to export orders. Check logs for details.', 'error')

        return redirect(url_for('database_view'))

    except Exception as e:
        logger.error(f"Export error: {e}")
        logger.error(traceback.format_exc())
        flash(f'Export failed: {str(e)}', 'error')
        return redirect(url_for('database_view'))


@app.route('/download-csv', methods=['GET'])
@error_handler
def download_csv():
    """Download all orders as CSV directly (no cloud upload)"""
    logger.info("DOWNLOAD-CSV route called")

    db = get_db_safe()
    if not db:
        flash('Database not available', 'error')
        return redirect(url_for('database_view'))

    try:
        # Get limit from query parameter
        limit = int(request.args.get('limit', 10000))

        # Get orders
        orders = db.get_recent_orders(limit=limit)

        if not orders:
            flash('No orders to export', 'warning')
            return redirect(url_for('database_view'))

        # Generate CSV in memory with GSM Fusion format
        output = io.StringIO()
        fieldnames = [
            'SERVICE', 'IMEI NO.', 'CREDITS', 'STATUS', 'CODE',
            'IMEI 2', 'CARRIER', 'SIMLOCK', 'MODEL', 'FMI',
            'ORDER DATE', 'NOTES'
        ]

        writer = csv.DictWriter(output, fieldnames=fieldnames, delimiter='\t')
        writer.writeheader()

        for order in orders:
            # Format credits with $ prefix
            credits = order.get('credits', '')
            if credits and str(credits).replace('.', '', 1).isdigit():
                credits = f"${credits}"

            row = {
                'SERVICE': order.get('service_name', ''),
                'IMEI NO.': order.get('imei', ''),
                'CREDITS': credits,
                'STATUS': order.get('status', ''),
                'CODE': order.get('result_code_display', '') or order.get('result_code', ''),
                'IMEI 2': order.get('imei2', ''),
                'CARRIER': order.get('carrier', ''),
                'SIMLOCK': order.get('simlock', ''),
                'MODEL': order.get('model', ''),
                'FMI': order.get('fmi', ''),
                'ORDER DATE': order.get('order_date', ''),
                'NOTES': order.get('notes', '')
            }
            writer.writerow(row)

        # Create response with CSV
        csv_data = output.getvalue()

        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'orders_export_{timestamp}.csv'

        logger.info(f"Generated CSV download: {len(orders)} orders, {len(csv_data)} bytes")

        return Response(
            csv_data,
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )

    except Exception as e:
        logger.error(f"CSV download error: {e}")
        logger.error(traceback.format_exc())
        flash(f'Download failed: {str(e)}', 'error')
        return redirect(url_for('database_view'))


@app.route('/download-completed-csv', methods=['GET'])
@error_handler
def download_completed_csv():
    """Download completed orders as CSV directly (no cloud upload)"""
    logger.info("DOWNLOAD-COMPLETED-CSV route called")

    db = get_db_safe()
    if not db:
        flash('Database not available', 'error')
        return redirect(url_for('database_view'))

    try:
        # Get completed orders
        orders = db.get_orders_by_status('Completed')

        if not orders:
            flash('No completed orders to export', 'warning')
            return redirect(url_for('database_view'))

        # Generate CSV in memory with GSM Fusion format
        output = io.StringIO()
        fieldnames = [
            'SERVICE', 'IMEI NO.', 'CREDITS', 'STATUS', 'CODE',
            'IMEI 2', 'CARRIER', 'SIMLOCK', 'MODEL', 'FMI',
            'ORDER DATE', 'NOTES'
        ]

        writer = csv.DictWriter(output, fieldnames=fieldnames, delimiter='\t')
        writer.writeheader()

        for order in orders:
            # Format credits with $ prefix
            credits = order.get('credits', '')
            if credits and str(credits).replace('.', '', 1).isdigit():
                credits = f"${credits}"

            row = {
                'SERVICE': order.get('service_name', ''),
                'IMEI NO.': order.get('imei', ''),
                'CREDITS': credits,
                'STATUS': order.get('status', ''),
                'CODE': order.get('result_code_display', '') or order.get('result_code', ''),
                'IMEI 2': order.get('imei2', ''),
                'CARRIER': order.get('carrier', ''),
                'SIMLOCK': order.get('simlock', ''),
                'MODEL': order.get('model', ''),
                'FMI': order.get('fmi', ''),
                'ORDER DATE': order.get('order_date', ''),
                'NOTES': order.get('notes', '')
            }
            writer.writerow(row)

        # Create response with CSV
        csv_data = output.getvalue()

        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'completed_orders_{timestamp}.csv'

        logger.info(f"Generated CSV download: {len(orders)} completed orders, {len(csv_data)} bytes")

        return Response(
            csv_data,
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )

    except Exception as e:
        logger.error(f"CSV download error: {e}")
        logger.error(traceback.format_exc())
        flash(f'Download failed: {str(e)}', 'error')
        return redirect(url_for('database_view'))


@app.route('/list-exports', methods=['GET'])
@error_handler
def list_exports():
    """List all exported CSV files from Supabase Storage"""
    logger.info("LIST-EXPORTS route called")

    try:
        # List exported files
        files = list_exported_csvs(limit=50)

        if files:
            # Format file list for display
            file_list = []
            for file in files:
                file_info = {
                    'name': file.get('name', 'Unknown'),
                    'size': file.get('size', 0),
                    'created_at': file.get('created_at', ''),
                    'url': file.get('url', '')
                }
                file_list.append(file_info)

            return jsonify({
                'success': True,
                'count': len(file_list),
                'files': file_list
            })
        else:
            return jsonify({
                'success': False,
                'message': 'No exported files found'
            })

    except Exception as e:
        logger.error(f"List exports error: {e}")
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.errorhandler(404)
def not_found(e):
    """Handle 404 errors"""
    return render_template('error.html', error="Page not found"), 404


@app.errorhandler(500)
def server_error(e):
    """Handle 500 errors"""
    logger.error(f"500 error: {e}")
    return render_template('error.html', error="Internal server error"), 500


@app.before_request
def before_request():
    """Log all requests for debugging"""
    logger.info(f"Request: {request.method} {request.path}")


@app.after_request
def after_request(response):
    """Log all responses"""
    logger.info(f"Response: {response.status_code} for {request.path}")
    return response


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5001))

    # Pre-warm cache on startup
    logger.info("Starting application...")
    logger.info("Pre-warming services cache...")
    try:
        services = get_services_cached()
        logger.info(f"✓ Cache warmed with {len(services)} services")
    except Exception as e:
        logger.error(f"Failed to warm cache: {e}")
        logger.warning("Continuing anyway - cache will populate on first request")

    logger.info(f"Starting Flask on port {port}")
    app.run(host='0.0.0.0', port=port, debug=False)
